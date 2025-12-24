
import os
import cv2
import numpy as np
import mmap
import contextlib

from PyQt5.QtGui import QPixmap, QImage
from matplotlib import cm
from matplotlib import pyplot as plt
 
from scipy import signal
from scipy.integrate import trapezoid
from scipy.signal import welch
from .Tint_Matlab import bits2uV
from ..data_loaders import get_output_filename

# =========================================================================== # 

def tot_power_in_fband(fft_xaxis: np.ndarray, fft_yaxis: np.ndarray, 
                       low_freq: float, high_freq: float) -> float: 
    
    '''
        Computes the total power within a specified frequency band of a signal

        Params:
            fft_axis (np.ndarray):
                fast fourier transform axis array of frequencies 
            fft_yaxis (np.ndarray):
                fast fourier transform amplitude array of signal
            low_freq, high_freq (float):
                Low and high end of frequency band respectively
        
        Returns:
            tot_power (np.float): Value of total power in specific band
    '''

    # Grab indices of values between freq band of interest in the fft
    indices = np.where(  ((fft_xaxis >= low_freq) & (fft_xaxis <= high_freq)) )[0]
    fft_xaxis_band = fft_xaxis[indices]
    fft_yaxis_band = fft_yaxis[indices]
    
    # Computes the total power in band via integration (trapezoidal rule)
    tot_power = trapezoid(fft_yaxis_band, x=fft_xaxis_band)
    
    return tot_power

# =========================================================================== # 

def tPowers_Fband_chunked(chunks: np.ndarray, fs: int, window: str, 
                          freq_low: float, freq_high: float) -> tuple:
    
    '''
        Computes the total power per chunk of a chunked signal within
        a specific frequency band

        Params: 
            chunks (np.ndarray): 
                List of numpy arrays representing the chunked signal
            fs (int): 
                Sampling frequency
            window (str): 
                Window type for pwelch (eg. hann, blackmann...)
            freq_low , freq_high (float): 
                Lower and upper bounds of specified frequency band of interest respectively
        
        Returns: 
            tuple: tot_chunk_pows, plot_data
            --------
            tot_chunk_pows (np.ndarray): 
                array of total powers per chunk for specific freq band
            plot_data (np.ndarray):
                array of power spectrum densities and frequencies per chunk
    '''
    
    # Instantiate empty arrays in preparation for plotting and 
    # consolidating total chunk powers
    plot_data = np.empty((len(chunks), 1), dtype=tuple)
    tot_chunk_pows = np.empty((len(chunks), 1), dtype=tuple)
    for i, data_chunk in enumerate(chunks): 
            # Compute the spectral density per chunk using welch method
            # Compute no. samples per segment for welch fft
            overlap = 0.5
            nPerSeg = len(data_chunk)
            nOverlap = np.round(overlap * nPerSeg)
            freq, psd = welch(data_chunk, fs=fs, window=window, 
                    nperseg=nPerSeg, noverlap=nOverlap, detrend="constant", 
                    return_onesided=True, scaling="density")
            
            # Compute average power within specified spectrum band per chunk
            plot_data[i][0] = (freq, psd) 
            tot_chunk_pows[i] = tot_power_in_fband(freq, psd, freq_low, freq_high)
            
    return tot_chunk_pows, plot_data
    
# =========================================================================== #

def compute_tracking_chunks(pos_x, pos_y, pos_t, chunk_size, n_chunks=None):
    
    '''
        Chunks the position data into 'n' second intervals based on chunk size.
        Used for plotting subject tracking

        Params:
            pos_x, pos_y, pos_t (np.ndarray) :
                x,y position data coordinates and timestamp arrays respectively
            chunk_size (int):
                Length of each chunk of position data in seconds
            n_chunks (int, optional):
                Expected number of chunks (aligns with EEG chunks to prevent off-by-one loss)

        Returns:
            Tuple: pos_x_chunks, pos_y_chunks
            --------
            pos_x_chunks (list):
                List of x coordinate data in the following format:
                    [ [x coordinates between 0s and n seconds], [x coordinates between 0 seconds and 2n seconds...] ...etc]
            pos_y_chunks (list):
                List of y coordinate data in the following format:
                    [ [y coordinates between 0s and n seconds], [y coordinates between 0 seconds and 2n seconds...] ...etc]
    '''

    # Align number of chunks to the EEG chunk count when available
    if n_chunks is None:
        n_chunks = int(np.ceil(float(pos_t[-1]) / float(chunk_size)))
    n_chunks = max(int(n_chunks), 1)

    # Build chunk boundaries in time, then convert to indices while ensuring full coverage
    spaced_t = np.arange(0, n_chunks + 1, dtype=float) * float(chunk_size)
    if spaced_t[-1] < float(pos_t[-1]):
        spaced_t[-1] = float(pos_t[-1])

    boundaries = np.searchsorted(pos_t, spaced_t, side='left')
    boundaries[-1] = len(pos_t)
    boundaries = np.clip(boundaries, 0, len(pos_t))
    boundaries = np.maximum.accumulate(boundaries)

    pos_x_chunks = []
    pos_y_chunks = []

    # Iterate through each 'time chunk', and append the tracking data per chunk (cumulative slices)
    for i in range(1, len(boundaries)):
        pos_x_chunks.append(pos_x[:boundaries[i]])
        pos_y_chunks.append(pos_y[:boundaries[i]])

    return pos_x_chunks, pos_y_chunks

# =========================================================================== #

def compute_freq_map(self, freq_range: str, pos_x: np.ndarray, pos_y: np.ndarray, pos_t: np.ndarray, 
                     fs: int, chunks: list, chunk_size: int, **kwargs) -> list:
    
    '''
        Maps the frequency power in a specific band as a function of subject position

        Params:
            freq_range (str):
                Choose frequency band ('Delta', 'Theta', 'Beta', 'Low_Gamma', or 'High_Gamma')
            pos_x, pos_y, pos_t (np.ndarray): 
                x,y coordinates, and timestamp arrays respectively
            fs (int):
                Sampling frequency
            chunks (list):
                List containing 'n' second signal chunks
            chunk_size (int):
                Length of signal chunks in an integer number of seconds

        Kwargs*:
            scaling_factor_perBand (dict):
                Dictionary containing the normalization factor for each frequency band
            chunk_pows_perBand (dict):
                Dictionary containing the total power across the whole signal in each band

        Returns:
            List: maps
            ------
            A list of frequency maps, one map per signal chunk. This is for one frequency band. 
    '''

    # Grab kwargs
    chunk_pows_perBand = kwargs.get('chunk_pows_perBand', None)
    scaling_factor_perBand = kwargs.get('scaling_factor_perBand', None)
    
    # Set kernel and std block sizes to about 20% of the map sizes in preparation for convolutional smoothing.
    kernlen = int(256*0.2)
    std = int(kernlen*0.2)

    # Smooth via kernel convolution
    kernel = gkern(kernlen,std)
    
    # Vector that will help associate each chunk with the current time in pos_t
    time_vec = np.linspace(0,pos_t[-1], len(chunks))
    
    ########### Compute the spectral density per chunk ###########
    # Min and max dimensions of arena for scaling 
    min_x = min(pos_x)
    max_x = max(pos_x)
    min_y = min(pos_y)
    max_y = max(pos_y)
    
    # Calculate the new dimensions of the frequncy map
    resize_ratio = np.abs(max_x - min_x) / np.abs(max_y - min_y)
    # Ensure the largest dimension of the map is 256
    base_resolution_scale = 256
    
    # Adjust the map dimensions based on which side is longer or shorter
    if resize_ratio > 1:
      x_resize = int(np.ceil (base_resolution_scale*resize_ratio)) 
      y_resize = base_resolution_scale
      
    elif resize_ratio < 1: 
        x_resize = base_resolution_scale
        y_resize = int(np.ceil (base_resolution_scale*(1/resize_ratio)))
     
    # If the map is perfectly symmetrical, set both sides to 256
    else: 
        x_resize = base_resolution_scale
        y_resize = base_resolution_scale
    
    # Equally part the x and y dimensions of the map into 256 spatial bins
    row_values = np.linspace(min_x,max_x,256)
    column_values = np.linspace(min_y,max_y,256)
    
    # Build chunk boundaries based on the EEG chunk count to avoid off-by-one drop
    n_chunks = len(chunks)
    spaced_t = np.arange(0, n_chunks + 1, dtype=float) * float(chunk_size)
    if spaced_t[-1] < float(pos_t[-1]):
        spaced_t[-1] = float(pos_t[-1])

    boundaries = np.searchsorted(pos_t, spaced_t, side='left')
    boundaries[-1] = len(pos_t)
    boundaries = np.clip(boundaries, 0, len(pos_t))
    boundaries = np.maximum.accumulate(boundaries)
    
    # Compute a scaling factor for the occupancy map
    occ_fs = float(pos_t[1] - pos_t[0])
    occ_scaling_factor = len(pos_t) * occ_fs
    maximum_value = index = 0
    # Initialize the number of maps based on the number of EEG chunks
    maps = [None] * n_chunks
    
    # Iterate through each 'time chunk', and generate the maps per chunk
    for i in range(n_chunks):
        start_idx = boundaries[i]
        end_idx = boundaries[i + 1]
        
        # Initialize empty occupancy and eeg maps
        occ_map_raw = np.zeros((256,256))
        eeg_map_raw = np.zeros((256,256))
        
        # Build the occupancy and eeg maps for the current time chunk
        for j in range(start_idx, end_idx):
            # Grab the row and column where the mice is based on timestamp
            row_index = np.abs(row_values - float(pos_x[j])).argmin()
            column_index = np.abs(column_values - float(pos_y[j])).argmin()
            
            # Encode the power of the band into whatever bins have been visited
            eeg_map_raw[row_index][column_index] += chunk_pows_perBand[freq_range][np.abs(time_vec - float(pos_t[j])).argmin()][0]
            # Encode the occupancy based on whatever bins have been visited
            occ_map_raw[row_index][column_index] += occ_fs
            
        # Normalize the maps using the scaling factors
        eeg_map_normalized = eeg_map_raw / (scaling_factor_perBand[freq_range])
        occ_map_normalized = occ_map_raw / occ_scaling_factor

        # Compute the frequency map by dividing the eeg map with the occupancy map.
        # This follows the same compute principle as a neural ratemap. Regions of high 
        # eeg activity and low occupancy will encode as significant activity.
        fMap = np.divide(
            eeg_map_normalized,
            occ_map_normalized,
            out=np.zeros_like(eeg_map_normalized),
            where=occ_map_normalized != 0,
        )
        # Replace any Nans with zero in case zero divide is encountered
        fMap[np.isnan(fMap)] = 0
        # Rotate and smooth the map
        fMap_smoothed = np.rot90(cv2.filter2D(fMap,-1,kernel))
        
        # Keep track of the largest value encountered in the sequence of smoothed maps
        if max(fMap_smoothed.flatten()) > maximum_value:
            maximum_value = max(fMap_smoothed.flatten())
        
        maps[index] = fMap_smoothed

        index+=1
    
    # Finally, divide all the maps by the largest value encountered. This ensures that 
    # all the maps scale between 0 and 1, which is needed for plt.imshow function later on.
    # This scaling preserves the relative relationships within each map, while
    # also normalizing the maps relative to eachother. (i.e we can deduce differences in
    # frequency strengh between the maps by comparing their brightness's to eachother.) 
    for i in range(len(maps)): 
        smoothed_and_normalized_map = (maps[i] / maximum_value)
        smoothed_and_normalized_map = cv2.resize(smoothed_and_normalized_map, dsize=(x_resize, y_resize), interpolation=cv2.INTER_CUBIC)
        rgba = np.uint8(cm.jet(smoothed_and_normalized_map) * 255)
        h, w, _ = rgba.shape
        qimage = QImage(rgba.data, w, h, 4 * w, QImage.Format_RGBA8888)
        maps[i] = QPixmap.fromImage(qimage)
        
    return maps
    
# =========================================================================== #

def compute_scaling_and_tPowers(self, window: str, pos_x: np.ndarray, pos_y: np.ndarray, 
                                pos_t: np.ndarray, chunks: list, fs: int) -> tuple:
    
    '''
        Compute 2 sets of scaling factors. The first set helps normalize maps belonging to the same band. 
        The second helps compute what percentage each bands total power contributes to the entire signal. 

        Params:
            window (str):
                Type of window for fourier transform. Choose between 'hamming','hann','backmanharris','boxcar'
            pos_x, pos_y, pos_t (np.ndarray): 
                x,y coordinates, and timestamp arrays respectively
            chunks (list):
                List of chunked signals, each 'n' seconds long
            fs (int):
                Sampling frequency for Welch method. 

        Return:
            Tuple: scaling_factor_perBand, scaling_factor_crossband, chunk_pows_perBand, plot_data
            --------
            scaling_factor_perBand (dict):
                Dictionary containing the total power of each band, indexed using band name.
            scaling_factor_crossband (dict):
                Dictionary containing the percentage contribution of each band to the total 
                signal power.
            chunk_pows_perBand (dict):
                Dictionary containing array of chunk powers per band
            plot_data (np.ndarray):
                Array of power spectrum densities and frequencies per chunk
    '''

    # Smooth via kernel convolution
    # Set kernel and stdev block sizes to 20% map size
    kernlen = int(256*0.2)
    std = int(kernlen*0.2)
    kernel = gkern(kernlen,std)
    
    # Create dictionary defining frequency bands and corresponding ranges in Hz
    freq_ranges = {'Delta': np.array([1, 3]), 
                   'Theta': np.array([4, 12]),
                   'Beta': np.array([13, 20]),
                   'Low Gamma': np.array([35, 55]),
                   'High Gamma': np.array([65, 120]),
                   'Ripple': np.array([80, 250]),
                   'Fast Ripple': np.array([250, 500])} #Abid
    
    # Dict to store total power per band
    scaling_factor_perBand = dict()
    # Vector that will help associate each chunk with the current time in pos_t
    time_vec = np.linspace(0,pos_t[-1], len(chunks))
    
    # Will store the total chunk powers per freq band
    chunk_pows_perBand = dict()
    
    ########### Compute the spectral density per chunk ###########
    
    # Grab chunks
    # Calculate a scaling factor by adding up all the power PER band. This will help us 
    # normalize the eeg graphs per band. 
    for key, value in freq_ranges.items():
        self.signals.text_progress.emit("Computing " + key + " scaling")
        # Compute chunk powers for each band
        chunk_pows, plot_data = tPowers_Fband_chunked(chunks, fs, window, value[0], value[1])
        # Using the number of samples in each chunk, compute the total power in each band. 
        nSamples_perChunk = (time_vec[1] - time_vec[0])/(1/fs)
        total_powInBand = sum(chunk_pows) * nSamples_perChunk
        
        chunk_pows_perBand[key] = chunk_pows
        scaling_factor_perBand[key] = total_powInBand[0]
        
        print(key + ' scaling is done')
    self.signals.text_progress.emit("Scaling complete!")    

    # Calculate a scaling factor as a proportion of the total power of all the bands in the signal
    # By this logic, Delta will make up x% of signal power, Beta = y% ...etc such that
    # the total ratio will add up to 100%. This will help us get a sense of relation between the signal bands
    sum_values = sum(scaling_factor_perBand.values())
    scaling_factor_crossband = dict()
    for key, value in freq_ranges.items():
        scaling_factor_crossband[key] = scaling_factor_perBand[key] / sum_values
    
    return scaling_factor_perBand, scaling_factor_crossband, chunk_pows_perBand, plot_data

# =========================================================================== #

def speed_bins(lower_speed: float, higher_speed: float, pos_v: np.ndarray, 
               pos_x: np.ndarray, pos_y: np.ndarray, pos_t: np.ndarray) -> tuple: 
    
    '''
        Selectively filters position values of subject travelling between
        specific speed limits. 
        
        Params: 
            lower_speed (float): 
                Lower speed bound (cm/s) 
            higher_speed (float): 
                Higher speed bound (cm/s) 
            pos_v (np.ndarray): 
                Array holding speed values of subject for entire experiment
            pos_x, pos_y, pos_t (np.ndarray): X, Y coordinates and timestamp array of subject respectively
            
            Returns: 
                Tuple: new_pos_x, new_pos_y, new_pos_t  
                --------
                New x,y coordinates and time array 
    '''
        
    chosen_indices = np.where((pos_v >= lower_speed) & (pos_v <= higher_speed))
    new_pos_x = pos_x[chosen_indices].flatten()
    new_pos_y = pos_y[chosen_indices].flatten()
    new_pos_t = pos_t[chosen_indices].flatten()
            
    return new_pos_x, new_pos_y, new_pos_t  

# =========================================================================== #  

# Credit:
# https://stackoverflow.com/questions/29731726/how-to-calculate-a-gaussian-kernel-matrix-efficiently-in-numpy

def gkern(kernlen, std):
    
    '''
        Returns a 2D Gaussian kernel array
    '''

    gkern1d = signal.windows.gaussian(kernlen, std=std).reshape(kernlen, 1)
    gkern2d = np.outer(gkern1d, gkern1d)
    return gkern2d

# =========================================================================== #

# Credit:
# https://stackoverflow.com/questions/44526121/finding-closest-values-in-two-numpy-arrays

def finder(a, b):

    '''
         Finds the common indices between two arrays
    '''
    dup = np.searchsorted(a, b)
    uni = np.unique(dup)
    uni = uni[uni < a.shape[0]]
    ret_b = np.zeros(uni.shape[0])
    for idx, val in enumerate(uni):
        bw = np.argmin(np.abs(a[val]-b[dup == val]))
        tt = dup == val
        ret_b[idx] = np.where(tt == True)[0][bw]
    
    common_indices = np.column_stack((uni, ret_b))
    return common_indices[:,0].astype(int)

# =========================================================================== #

def compute_binned_freq_analysis(pos_x: np.ndarray, pos_y: np.ndarray, pos_t: np.ndarray, 
                                 fs: int, chunks: list, chunk_size: int, **kwargs) -> dict:
    
    '''
        Divides frequency maps into 4x4 spatial bins and analyzes power distribution 
        across multiple frequency bands. Tracks changes across time chunks.

        Params:
            pos_x, pos_y, pos_t (np.ndarray): 
                x,y coordinates, and timestamp arrays respectively
            fs (int):
                Sampling frequency
            chunks (list):
                List containing 'n' second signal chunks
            chunk_size (int):
                Length of signal chunks in an integer number of seconds

        Kwargs:
            chunk_pows_perBand (dict):
                Dictionary containing the total power per chunk for each frequency band
            scaling_factor_perBand (dict, optional):
                Dictionary containing normalization factors per band (not required)

        Returns:
            dict: binned_analysis_data
            ------
            Dictionary structure:
            {
                'bins_grid': 4x4 array of bin coordinates,
                'time_chunks': number of time chunks analyzed,
                'bands': list of frequency bands,
                'bin_power_timeseries': {band_name: 4x4x(n_chunks) array},
                'bin_dominant_band': {chunk_idx: 4x4 array of dominant band names},
                'bin_occupancy': 4x4 array of occupancy per bin,
                'bin_statistics': {band_name: {'mean': value, 'max': value, 'min': value, 'std': value}}
            }
    '''
    
    chunk_pows_perBand = kwargs.get('chunk_pows_perBand', None)
    scaling_factor_perBand = kwargs.get('scaling_factor_perBand', None)
    
    if chunk_pows_perBand is None:
        raise ValueError("chunk_pows_perBand required in kwargs")
    
    # Define frequency bands
    freq_ranges = {'Delta': np.array([1, 3]), 
                   'Theta': np.array([4, 12]),
                   'Beta': np.array([13, 20]),
                   'Low Gamma': np.array([35, 55]),
                   'High Gamma': np.array([65, 120]),
                   'Ripple': np.array([80, 250]),
                   'Fast Ripple': np.array([250, 500])}
    
    # Position boundaries
    min_x, max_x = min(pos_x), max(pos_x)
    min_y, max_y = min(pos_y), max(pos_y)
    
    # Create 4x4 bin boundaries
    x_bin_edges = np.linspace(min_x, max_x, 5)  # 5 edges for 4 bins
    y_bin_edges = np.linspace(min_y, max_y, 5)
    
    # Time vector for associating chunks
    n_time_chunks = len(chunks)
    time_vec = np.linspace(0, pos_t[-1], n_time_chunks)

    # Create time chunk indices aligned to EEG chunks (prevents dropping the last bin)
    spaced_t = np.arange(0, n_time_chunks + 1, dtype=float) * float(chunk_size)
    if spaced_t[-1] < float(pos_t[-1]):
        spaced_t[-1] = float(pos_t[-1])

    boundaries = np.searchsorted(pos_t, spaced_t, side='left')
    boundaries[-1] = len(pos_t)
    boundaries = np.clip(boundaries, 0, len(pos_t))
    boundaries = np.maximum.accumulate(boundaries)
    
    # Initialize data structures
    bin_power_timeseries = {band: np.zeros((4, 4, n_time_chunks)) 
                           for band in freq_ranges.keys()}
    bin_occupancy = np.zeros((4, 4))
    bin_dominant_band = [np.empty((4, 4), dtype=object) for _ in range(n_time_chunks)]
    
    # Compute scaling factor for occupancy
    occ_fs = float(pos_t[1] - pos_t[0])
    occ_scaling_factor = len(pos_t) * occ_fs
    
    # Process each time chunk
    for chunk_idx in range(n_time_chunks):
        chunk_num = chunk_idx
        start_idx = boundaries[chunk_idx]
        end_idx = boundaries[chunk_idx + 1]
        
        # Process each position sample in this time chunk
        for pos_idx in range(start_idx, end_idx):
            x_pos = float(pos_x[pos_idx])
            y_pos = float(pos_y[pos_idx])
            t_pos = float(pos_t[pos_idx])
            
            # Find which bin this position falls into
            x_bin = np.digitize(x_pos, x_bin_edges) - 1
            y_bin = np.digitize(y_pos, y_bin_edges) - 1
            
            # Clamp to valid bin range
            x_bin = np.clip(x_bin, 0, 3)
            y_bin = np.clip(y_bin, 0, 3)
            
            # Find closest chunk index
            chunk_time_idx = np.abs(time_vec - t_pos).argmin()
            
            # Accumulate power for each band in this bin
            for band_name in freq_ranges.keys():
                power_val = chunk_pows_perBand[band_name][chunk_time_idx][0]
                bin_power_timeseries[band_name][x_bin, y_bin, chunk_num] += power_val
            
            # Accumulate occupancy
            bin_occupancy[x_bin, y_bin] += occ_fs
        
        # Determine dominant band per bin for this time chunk
        for x_bin in range(4):
            for y_bin in range(4):
                # Get power values for all bands in this bin at this time
                band_powers = {band: bin_power_timeseries[band][x_bin, y_bin, chunk_num]
                             for band in freq_ranges.keys()}
                
                # Find dominant band (highest power)
                dominant = max(band_powers, key=band_powers.get)
                bin_dominant_band[chunk_num][x_bin, y_bin] = dominant
    
    # Normalize occupancy-based power (per unit time spent in bin)
    normalized_bin_power = {}
    for band in freq_ranges.keys():
        normalized_bin_power[band] = np.zeros_like(bin_power_timeseries[band])
        for t in range(n_time_chunks):
            for x in range(4):
                for y in range(4):
                    if bin_occupancy[x, y] > 0:
                        normalized_bin_power[band][x, y, t] = (
                            bin_power_timeseries[band][x, y, t] / bin_occupancy[x, y]
                        )
    
    # Compute statistics across all time chunks for each band and bin
    bin_statistics = {}
    for band in freq_ranges.keys():
        bin_statistics[band] = {
            'mean': np.mean(normalized_bin_power[band], axis=2),
            'max': np.max(normalized_bin_power[band], axis=2),
            'min': np.min(normalized_bin_power[band], axis=2),
            'std': np.std(normalized_bin_power[band], axis=2)
        }
    
    # Package results
    binned_analysis_data = {
        'x_bin_edges': x_bin_edges,
        'y_bin_edges': y_bin_edges,
        'time_chunks': n_time_chunks,
        'bands': list(freq_ranges.keys()),
        'bin_power_timeseries': normalized_bin_power,
        'bin_dominant_band': bin_dominant_band,
        'bin_occupancy': bin_occupancy,
        'bin_statistics': bin_statistics
    }
    
    return binned_analysis_data

# =========================================================================== #

def export_binned_analysis_to_csv(binned_data: dict, output_path: str):
    
    '''
        Exports binned frequency analysis to Excel files with multiple sheets per metric

        Params:
            binned_data (dict):
                Dictionary returned from compute_binned_freq_analysis
            output_path (str):
                Base path for output files (without extension)
    '''
    
    # Prefer Excel via openpyxl if available; else fall back to CSV
    try:
        import openpyxl
        use_excel = True
    except ImportError:
        use_excel = False
        import csv
    
    bands = binned_data['bands']
    
    # Precompute percent power across bands: (band power / total power per bin per chunk) * 100
    percent_power = {band: np.zeros_like(binned_data['bin_power_timeseries'][band]) for band in bands}
    n_chunks = binned_data['time_chunks']
    for t in range(n_chunks):
        total_power_chunk = np.zeros((4, 4))
        for band in bands:
            total_power_chunk += binned_data['bin_power_timeseries'][band][:, :, t]
        for band in bands:
            band_power_chunk = binned_data['bin_power_timeseries'][band][:, :, t]
            with np.errstate(divide='ignore', invalid='ignore'):
                pct = np.where(total_power_chunk > 0, (band_power_chunk / total_power_chunk) * 100.0, 0.0)
            percent_power[band][:, :, t] = pct

    # Percent power mean across time chunks
    percent_power_mean = {band: np.mean(percent_power[band], axis=2) for band in bands}

    if use_excel:
        # Mean power workbook
        mean_power_file = f"{output_path}_mean_power.xlsx"
        wb_mean = openpyxl.Workbook()
        # Remove default sheet
        wb_mean.remove(wb_mean.active)
        for band in bands:
            ws = wb_mean.create_sheet(title=band)
            mean_power = binned_data['bin_statistics'][band]['mean']
            for row in mean_power:
                ws.append([float(val) for val in row])
        wb_mean.save(mean_power_file)

        # Percent power workbook (mean percent per bin)
        percent_file = f"{output_path}_percent_power.xlsx"
        wb_pct = openpyxl.Workbook()
        wb_pct.remove(wb_pct.active)
        for band in bands:
            ws = wb_pct.create_sheet(title=band)
            mean_pct = percent_power_mean[band]
            for row in mean_pct:
                ws.append([float(val) for val in row])
        wb_pct.save(percent_file)

        # Occupancy workbook
        occ_file = f"{output_path}_occupancy.xlsx"
        wb_occ = openpyxl.Workbook()
        ws_occ = wb_occ.active
        ws_occ.title = 'Occupancy'
        occ = binned_data['bin_occupancy']
        for row in occ:
            ws_occ.append([float(val) for val in row])
        wb_occ.save(occ_file)

        # Dominant band counts workbook (one sheet per band)
        dominant_file = f"{output_path}_dominant_band.xlsx"
        wb_dom = openpyxl.Workbook()
        wb_dom.remove(wb_dom.active)
        dominant_counts = {band: np.zeros((4, 4)) for band in bands}
        for chunk_data in binned_data['bin_dominant_band']:
            for x in range(4):
                for y in range(4):
                    band = chunk_data[x, y]
                    dominant_counts[band][x, y] += 1
        for band in bands:
            ws = wb_dom.create_sheet(title=band)
            counts = dominant_counts[band]
            for row in counts:
                ws.append([float(val) for val in row])
        wb_dom.save(dominant_file)

        return {
            'format': 'excel',
            'files': [mean_power_file, percent_file, occ_file, dominant_file],
            'reason': None
        }
    
    else:
        # Fallback to CSV format
        csv_files = []
        for band in binned_data['bands']:
            mean_power = binned_data['bin_statistics'][band]['mean']
            out_file = f"{output_path}_binned_{band}_mean_power.csv"
            with open(out_file, 'w', newline='') as f:
                writer = csv.writer(f)
                for row in mean_power:
                    writer.writerow(row.tolist())
            csv_files.append(out_file)

        # Percent power mean per band
        for band in bands:
            mean_pct = percent_power_mean[band]
            out_file = f"{output_path}_binned_{band}_percent_power.csv"
            with open(out_file, 'w', newline='') as f:
                writer = csv.writer(f)
                for row in mean_pct:
                    writer.writerow(row.tolist())
            csv_files.append(out_file)
        
        occ_out = f"{output_path}_bin_occupancy.csv"
        occ = binned_data['bin_occupancy']
        with open(occ_out, 'w', newline='') as f:
            writer = csv.writer(f)
            for row in occ:
                writer.writerow(row.tolist())
        csv_files.append(occ_out)
        
        dominant_counts = {band: np.zeros((4, 4)) for band in bands}
        for chunk_data in binned_data['bin_dominant_band']:
            for x in range(4):
                for y in range(4):
                    band = chunk_data[x, y]
                    dominant_counts[band][x, y] += 1
        
        for band in bands:
            out_file = f"{output_path}_bin_dominant_{band}_frequency.csv"
            counts = dominant_counts[band]
            with open(out_file, 'w', newline='') as f:
                writer = csv.writer(f)
                for row in counts:
                    writer.writerow(row.tolist())
            csv_files.append(out_file)

        return {
            'format': 'csv',
            'files': csv_files,
            'reason': 'openpyxl_not_installed'
        }

# =========================================================================== #

def visualize_binned_analysis(binned_data: dict, save_path: str = None) -> None:
    
    '''
        Creates visualization of 4x4 binned frequency analysis

        Params:
            binned_data (dict):
                Dictionary returned from compute_binned_freq_analysis
            save_path (str, optional):
                Path to save visualization. If None, displays to screen
    '''
    
    fig, axes = plt.subplots(2, 4, figsize=(16, 8))
    fig.suptitle('4x4 Spatial Bins - Frequency Band Analysis', fontsize=14, fontweight='bold')
    
    bands = binned_data['bands']
    
    # First row: Mean power per band
    for idx, band in enumerate(bands[:4]):
        mean_power = binned_data['bin_statistics'][band]['mean']
        im = axes[0, idx].imshow(mean_power, cmap='hot', aspect='auto')
        axes[0, idx].set_title(f'{band}\n(Mean Power)')
        axes[0, idx].set_xticks([0, 1, 2, 3])
        axes[0, idx].set_yticks([0, 1, 2, 3])
        axes[0, idx].grid(True, alpha=0.3)
        plt.colorbar(im, ax=axes[0, idx])
    
    # Second row: Remaining bands
    remaining_bands = bands[4:]
    for idx, band in enumerate(remaining_bands):
        mean_power = binned_data['bin_statistics'][band]['mean']
        im = axes[1, idx].imshow(mean_power, cmap='hot', aspect='auto')
        axes[1, idx].set_title(f'{band}\n(Mean Power)')
        axes[1, idx].set_xticks([0, 1, 2, 3])
        axes[1, idx].set_yticks([0, 1, 2, 3])
        axes[1, idx].grid(True, alpha=0.3)
        plt.colorbar(im, ax=axes[1, idx])
    
    # Hide unused subplots
    for idx in range(len(remaining_bands), 4):
        axes[1, idx].axis('off')
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"Visualization saved to {save_path}")
    else:
        plt.show()
    
    plt.close()

# =========================================================================== #

def visualize_binned_analysis_by_chunk(binned_data: dict, chunk_idx: int, save_path: str = None, 
                                        show_percent: bool = False) -> None:
    
    '''
        Creates visualization of 4x4 binned frequency analysis for a specific time chunk

        Params:
            binned_data (dict):
                Dictionary returned from compute_binned_freq_analysis
            chunk_idx (int):
                Index of the time chunk to visualize (0-based)
            save_path (str, optional):
                Path to save visualization. If None, displays to screen
            show_percent (bool):
                If True, show percent power; if False, show absolute power
    '''
    
    # Validate chunk index
    n_chunks = binned_data['time_chunks']
    if chunk_idx < 0 or chunk_idx >= n_chunks:
        chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
    
    fig, axes = plt.subplots(2, 4, figsize=(16, 8))
    power_type = "Percent Power" if show_percent else "Power"
    fig.suptitle(f'4x4 Spatial Bins - Chunk {chunk_idx} (Frequency Band {power_type})', 
                 fontsize=14, fontweight='bold')
    
    bands = binned_data['bands']
    
    # Get min/max across all chunks for consistent color scale
    vmin_all = {}
    vmax_all = {}
    for band in bands:
        timeseries_data = binned_data['bin_power_timeseries'][band]
        
        if show_percent:
            # Convert to percent: power per band / total power per bin per chunk * 100
            percent_data = np.zeros_like(timeseries_data)
            for t in range(timeseries_data.shape[2]):
                for x in range(4):
                    for y in range(4):
                        total_power = sum(binned_data['bin_power_timeseries'][b][x, y, t] 
                                        for b in bands)
                        if total_power > 0:
                            percent_data[x, y, t] = (timeseries_data[x, y, t] / total_power) * 100
            vmin_all[band] = np.nanmin(percent_data)
            vmax_all[band] = np.nanmax(percent_data)
        else:
            vmin_all[band] = np.nanmin(timeseries_data)
            vmax_all[band] = np.nanmax(timeseries_data)
    
    # First row: First 4 bands
    for idx, band in enumerate(bands[:4]):
        chunk_power = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
        
        if show_percent:
            # Compute percent for this band
            total_power = sum(binned_data['bin_power_timeseries'][b][:, :, chunk_idx] 
                            for b in bands)
            chunk_power = np.divide(chunk_power, total_power, where=total_power>0, out=np.zeros_like(chunk_power)) * 100
        
        im = axes[0, idx].imshow(chunk_power, cmap='hot', aspect='auto',
                                 vmin=vmin_all[band], vmax=vmax_all[band])
        axes[0, idx].set_title(f'{band}')
        axes[0, idx].set_xticks([0, 1, 2, 3])
        axes[0, idx].set_yticks([0, 1, 2, 3])
        axes[0, idx].grid(True, alpha=0.3)
        cbar = plt.colorbar(im, ax=axes[0, idx])
        cbar.set_label('%' if show_percent else 'Power', fontsize=9)
    
    # Second row: Remaining bands
    remaining_bands = bands[4:]
    for idx, band in enumerate(remaining_bands):
        chunk_power = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
        
        if show_percent:
            # Compute percent for this band
            total_power = sum(binned_data['bin_power_timeseries'][b][:, :, chunk_idx] 
                            for b in bands)
            chunk_power = np.divide(chunk_power, total_power, where=total_power>0, out=np.zeros_like(chunk_power)) * 100
        
        im = axes[1, idx].imshow(chunk_power, cmap='hot', aspect='auto',
                                 vmin=vmin_all[band], vmax=vmax_all[band])
        axes[1, idx].set_title(f'{band}')
        axes[1, idx].set_xticks([0, 1, 2, 3])
        axes[1, idx].set_yticks([0, 1, 2, 3])
        axes[1, idx].grid(True, alpha=0.3)
        cbar = plt.colorbar(im, ax=axes[1, idx])
        cbar.set_label('%' if show_percent else 'Power', fontsize=9)
    
    # Hide unused subplots
    for idx in range(len(remaining_bands), 4):
        axes[1, idx].axis('off')
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    else:
        plt.show()
    
    plt.close()

# =========================================================================== #

def visualize_binned_occupancy_and_dominant(binned_data: dict, chunk_idx: int = None, 
                                             save_path: str = None) -> None:
    
    '''
        Creates visualization of bin occupancy and dominant band for time chunks

        Params:
            binned_data (dict):
                Dictionary returned from compute_binned_freq_analysis
            chunk_idx (int, optional):
                Index of specific time chunk. If None, shows aggregated occupancy 
                and dominant band counts across all chunks.
            save_path (str, optional):
                Path to save visualization.
    '''
    
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    
    n_chunks = binned_data['time_chunks']
    if chunk_idx is not None:
        if chunk_idx < 0 or chunk_idx >= n_chunks:
            chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
    
    # Left panel: Occupancy
    occupancy = binned_data['bin_occupancy']
    im1 = axes[0].imshow(occupancy, cmap='viridis', aspect='auto')
    axes[0].set_title('Bin Occupancy (Total Time Spent)')
    axes[0].set_xticks([0, 1, 2, 3])
    axes[0].set_yticks([0, 1, 2, 3])
    axes[0].grid(True, alpha=0.3)
    cbar1 = plt.colorbar(im1, ax=axes[0])
    cbar1.set_label('Occupancy (samples)', fontsize=9)
    
    # Right panel: Dominant band
    if chunk_idx is not None:
        # Show dominant band for specific chunk
        dominant_chunk = binned_data['bin_dominant_band'][chunk_idx]
        title = f'Dominant Band - Chunk {chunk_idx}'
        # Create numeric mapping for visualization
        bands = binned_data['bands']
        band_map = {band: idx for idx, band in enumerate(bands)}
        numeric_dominant = np.zeros((4, 4))
        for x in range(4):
            for y in range(4):
                band = dominant_chunk[x, y]
                numeric_dominant[x, y] = band_map.get(band, 0)
        
        im2 = axes[1].imshow(numeric_dominant, cmap='tab10', aspect='auto', vmin=0, vmax=len(bands)-1)
        axes[1].set_title(title)
    else:
        # Show dominant band frequency across all chunks
        bands = binned_data['bands']
        band_map = {band: idx for idx, band in enumerate(bands)}
        dominant_counts = {band: np.zeros((4, 4)) for band in bands}
        
        for chunk_data in binned_data['bin_dominant_band']:
            for x in range(4):
                for y in range(4):
                    band = chunk_data[x, y]
                    dominant_counts[band][x, y] += 1
        
        # Show the most dominant band per bin
        max_dominant = np.zeros((4, 4))
        for x in range(4):
            for y in range(4):
                max_band = max(dominant_counts.keys(), 
                             key=lambda b: dominant_counts[b][x, y])
                max_dominant[x, y] = band_map[max_band]
        
        im2 = axes[1].imshow(max_dominant, cmap='tab10', aspect='auto', vmin=0, vmax=len(bands)-1)
        axes[1].set_title('Most Frequent Dominant Band (Across All Chunks)')
    
    axes[1].set_xticks([0, 1, 2, 3])
    axes[1].set_yticks([0, 1, 2, 3])
    axes[1].grid(True, alpha=0.3)
    
    # Add colorbar with band labels
    bands = binned_data['bands']
    cbar2 = plt.colorbar(im2, ax=axes[1], ticks=range(len(bands)))
    cbar2.set_ticklabels(bands, fontsize=8)
    cbar2.set_label('Band', fontsize=9)
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    else:
        plt.show()
    
    plt.close()

# =========================================================================== #

def export_binned_analysis_jpgs(binned_data: dict, output_folder: str, base_name: str) -> int:
    
    '''
        Export all per-chunk JPG visualizations for binned analysis (batch mode).
        Exports mean power, percent power, dominant band per chunk, and occupancy once.

        Params:
            binned_data (dict):
                Dictionary returned from compute_binned_freq_analysis
            output_folder (str):
                Output folder for JPG files
            base_name (str):
                Base filename (without extension)
        
        Returns:
            int: Number of JPGs exported
    '''
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder, exist_ok=True)
    
    bands = binned_data['bands']
    n_chunks = binned_data['time_chunks']
    export_count = 0
    
    # Export mean power for all chunks (JPG, quality 85)
    for chunk_idx in range(n_chunks):
        fig, axes = plt.subplots(2, 4, figsize=(16, 8))
        fig.suptitle(f'4x4 Spatial Bins - Chunk {chunk_idx} (Frequency Band Power)', 
                     fontsize=14, fontweight='bold')
        
        # Get min/max across all chunks for consistent color scale
        vmin_all = {}
        vmax_all = {}
        for band in bands:
            timeseries_data = binned_data['bin_power_timeseries'][band]
            vmin_all[band] = np.min(timeseries_data)
            vmax_all[band] = np.max(timeseries_data)
        
        # First row: First 4 bands
        for idx, band in enumerate(bands[:4]):
            chunk_power = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            im = axes[0, idx].imshow(chunk_power, cmap='hot', aspect='auto',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[0, idx].set_title(f'{band}')
            axes[0, idx].set_xticks([0, 1, 2, 3])
            axes[0, idx].set_yticks([0, 1, 2, 3])
            axes[0, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[0, idx])
            cbar.set_label('Power', fontsize=9)
        
        # Second row: Remaining bands
        remaining_bands = bands[4:]
        for idx, band in enumerate(remaining_bands):
            chunk_power = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            im = axes[1, idx].imshow(chunk_power, cmap='hot', aspect='auto',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[1, idx].set_title(f'{band}')
            axes[1, idx].set_xticks([0, 1, 2, 3])
            axes[1, idx].set_yticks([0, 1, 2, 3])
            axes[1, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[1, idx])
            cbar.set_label('Power', fontsize=9)
        
        # Hide unused subplots
        for idx in range(len(remaining_bands), 4):
            axes[1, idx].axis('off')
        
        plt.tight_layout()
        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx}_mean_power.jpg")
        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
        plt.close(fig)
        export_count += 1
    
    # Export percent power for all chunks (JPG, quality 85)
    for chunk_idx in range(n_chunks):
        fig, axes = plt.subplots(2, 4, figsize=(16, 8))
        fig.suptitle(f'4x4 Spatial Bins - Chunk {chunk_idx} (Frequency Band Percent Power)', 
                     fontsize=14, fontweight='bold')
        
        # Precompute percent power for this chunk
        percent_power_chunk = {}
        total_power_chunk = np.zeros((4, 4))
        for band in bands:
            total_power_chunk += binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
        
        for band in bands:
            band_power_chunk = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            with np.errstate(divide='ignore', invalid='ignore'):
                pct = np.where(total_power_chunk > 0, (band_power_chunk / total_power_chunk) * 100.0, 0.0)
            percent_power_chunk[band] = pct
        
        # Get min/max for percent across all chunks
        vmin_all = {}
        vmax_all = {}
        for band in bands:
            vmin_all[band] = 0
            vmax_all[band] = 100
        
        # First row: First 4 bands
        for idx, band in enumerate(bands[:4]):
            im = axes[0, idx].imshow(percent_power_chunk[band], cmap='hot', aspect='auto',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[0, idx].set_title(f'{band}')
            axes[0, idx].set_xticks([0, 1, 2, 3])
            axes[0, idx].set_yticks([0, 1, 2, 3])
            axes[0, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[0, idx])
            cbar.set_label('%', fontsize=9)
        
        # Second row: Remaining bands
        remaining_bands = bands[4:]
        for idx, band in enumerate(remaining_bands):
            im = axes[1, idx].imshow(percent_power_chunk[band], cmap='hot', aspect='auto',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[1, idx].set_title(f'{band}')
            axes[1, idx].set_xticks([0, 1, 2, 3])
            axes[1, idx].set_yticks([0, 1, 2, 3])
            axes[1, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[1, idx])
            cbar.set_label('%', fontsize=9)
        
        # Hide unused subplots
        for idx in range(len(remaining_bands), 4):
            axes[1, idx].axis('off')
        
        plt.tight_layout()
        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx}_percent_power.jpg")
        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
        plt.close(fig)
        export_count += 1
    
    # Export occupancy only once
    fig_occ = plt.figure(figsize=(6, 5))
    ax = fig_occ.add_subplot(111)
    occ = binned_data['bin_occupancy']
    im = ax.imshow(occ, cmap='viridis', aspect='auto')
    ax.set_title('Bin Occupancy (Total Time Spent)', fontsize=12, fontweight='bold')
    ax.set_xticks([0, 1, 2, 3])
    ax.set_yticks([0, 1, 2, 3])
    ax.grid(True, alpha=0.3)
    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label('Occupancy (samples)', fontsize=9)
    jpg_path = os.path.join(output_folder, f"{base_name}_occupancy.jpg")
    fig_occ.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
    plt.close(fig_occ)
    export_count += 1
    
    # Export dominant band per chunk
    for chunk_idx in range(n_chunks):
        fig, ax = plt.subplots(figsize=(6, 5))
        
        dominant_chunk = binned_data['bin_dominant_band'][chunk_idx]
        band_map = {band: idx for idx, band in enumerate(bands)}
        numeric_dominant = np.zeros((4, 4))
        for x in range(4):
            for y in range(4):
                band = dominant_chunk[x, y]
                numeric_dominant[x, y] = band_map.get(band, 0)
        
        im = ax.imshow(numeric_dominant, cmap='tab10', aspect='auto', vmin=0, vmax=len(bands)-1)
        ax.set_title(f'Dominant Band - Chunk {chunk_idx}', fontsize=12, fontweight='bold')
        ax.set_xticks([0, 1, 2, 3])
        ax.set_yticks([0, 1, 2, 3])
        ax.grid(True, alpha=0.3)
        
        cbar = plt.colorbar(im, ax=ax, ticks=range(len(bands)))
        cbar.set_ticklabels(bands, fontsize=8)
        cbar.set_label('Band', fontsize=9)
        
        plt.tight_layout()
        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx}_dominant_band.jpg")
        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
        plt.close(fig)
        export_count += 1
    
    return export_count



