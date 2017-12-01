import matplotlib
matplotlib.use('QT4Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
import os, sys, datetime
from scipy import misc, ndimage
import cv2
from scipy.io import savemat, loadmat
from skimage import color, io
import matplotlib.image as img
import skimage.measure
from core.Tint_Matlab import *
import numpy as np
import scipy.signal
from PyQt4 import QtGui, QtCore
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def QuadrantAnalysis(self, set_directory, arena, arena_shape=None):

    if arena_shape is not None:
        if arena_shape.lower() == 'circle':
            circle = True
        else:
            circle = False

    plot_linewidth = 5.334
    debug_figs = False
    # puts a QuadrantAnalysis directory in the set_directory
    save_directory = os.path.join(set_directory, 'QuadrantAnalysis')

    # adds a Figures directory to that QuadrantAnalysis directory to save the figures
    save_figures_directory = os.path.join(save_directory, 'Figures')

    # the name of the directory that the position files are located in (used for the excel filename)
    directory_name = os.path.basename(os.path.dirname(set_directory))

    # make the figure if it doesn't exist already
    if not os.path.exists(save_figures_directory):
        os.makedirs(save_figures_directory)

    set_directory_flist = os.listdir(set_directory)

    setfile_list = [os.path.join(set_directory, file)
                     for file in set_directory_flist if '.set' in file]

    if len(setfile_list) == 0:
        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: There are no .set files in the following directory: %s' % (str(datetime.datetime.now().date()),
                                                           str(datetime.datetime.now().time())[:8], set_directory))
        self.stop_analysis()
        return

    for set_filename in setfile_list:

        session_path = os.path.dirname(set_filename)  # name of the directory that the set file exists in

        session = os.path.splitext(os.path.basename(set_filename))[0]  # get the session name (basename of .set file)

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: analyzing the following session: %s!' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8], session))

        # creates the excel filename
        excel_filename = os.path.join(save_directory, '%s_quadrant.xlsx' % session)

        if os.path.exists(excel_filename):
            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: The following excel filename has already been analyzed, skipping: %s' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8], excel_filename))
            continue  # the file has already been analyzed

        posfile = os.path.join(session_path, session + '.pos')  # defining the position filename

        if not os.path.exists(posfile):
            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: the following position file does not exist, skipping: %s.' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8], posfile))
            continue

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: Loading position data.' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8]))

        posx, posy, post = getpos(posfile, arena)  # getting the mouse position

        # centering the positions
        center = centerBox(posx, posy)
        posx = posx - center[0]
        posy = posy - center[1]

        # Threshold for how far a mouse can move (100cm/s), in one sample (sampFreq = 50 Hz
        threshold = 100 / 50  # defining the threshold

        posx, posy, post = remBadTrack(posx, posy, post, threshold)  # removing bad tracks (faster than threshold)

        nonNanValues = np.where(np.isnan(posx) == False)[0]
        # removeing any NaNs
        post = post[nonNanValues]
        posx = posx[nonNanValues]
        posy = posy[nonNanValues]

        if len(posx) == 0:
            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: There are no valid positions (all NaNs)' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))
            continue

        # smooth positions with sgolay filter
        # posx = scipy.signal.savgol_filter(posx.flatten(), 15, 4)  # window of 15, order of 4
        # posy = scipy.signal.savgol_filter(posy.flatten(), 15, 4)

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: smoothing the position data' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8]))

        for index in range(7, len(posx) - 7):
            if index == len(posx) - 8:
                posx[index] = np.nanmean(posx[index - 7:])
                posy[index] = np.nanmean(posy[index - 7:])
            else:
                posx[index] = np.nanmean(posx[index - 7:index + 7 + 1])
                posy[index] = np.nanmean(posy[index - 7:index + 7 + 1])

        # calculating if the arena is a circle or not

        points = np.hstack((posx.reshape((len(posx), 1)),
                            posy.reshape((len(posy), 1))))  # making an array [x1, y1;x2,y2,...]

        dimensions = np.array([np.amin(posx), np.amax(posx),
                               np.amin(posy), np.amax(posy)])

        if arena_shape is None:
            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: determining the shape of the arena!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))

            # check if it is a square or a circle arena by looking at the corners
            bin_x = np.linspace(np.amin(posx), np.amax(posx), 11)
            bin_y = np.linspace(np.amin(posx), np.amax(posx), 11)

            corners = np.array([[bin_x[-2], bin_y[-2]], [bin_x[1], bin_y[-2]],
                                [bin_x[1], bin_y[1]], [bin_x[-2], bin_y[1]]])

            plot_corners = False
            if plot_corners:
                # plots to see the test if it is a square or cirlce. I essentially
                # determine if there is data at the corners of the arena
                fig_corners = plt.figure()
                ax_corners = fig_corners.add_subplot(111)
                ax_corners.plot(posx, posy, 'r-')
                ax_corners.plot(corners[0:2, 0], corners[0:2, 1], 'g')
                ax_corners.plot(corners[1:3, 0], corners[1:3, 1], 'g')
                ax_corners.plot(corners[[0, 3], 0], corners[[0, 3], 1], 'g')
                ax_corners.plot(corners[[2, 3], 0], corners[[2, 3], 1], 'g')

            circle_bool = []
            for corner in range(4):
                # for each corner, see if there is data in the corners
                if corner == 0:  # NE corner
                    bool_val = (points[:, 0] >= 0) * (points[:, 1] >= 0)
                    current_points = points[bool_val, :]

                    circle_bool.append(np.sum((current_points[:, 0] >= corners[corner, 0]) *
                                              (current_points[:, 1] >= corners[corner, 1])))

                elif corner == 1:  # NW Corner
                    bool_val = (points[:, 0] < 0) * (points[:, 1] >= 0)
                    current_points = points[bool_val, :]

                    circle_bool.append(np.sum((current_points[:, 0] < corners[corner, 0]) *
                                              (current_points[:, 1] >= corners[corner, 1])))

                elif corner == 2:  # SW Corner
                    bool_val = (points[:, 0] < 0) * (points[:, 1] < 0)
                    current_points = points[bool_val, :]

                    circle_bool.append(np.sum((current_points[:, 0] <= corners[corner, 0]) *
                                              (current_points[:, 1] < corners[corner, 1])))

                else:  # SE corner
                    bool_val = (points[:, 0] > 0) * (points[:, 1] < 0)
                    current_points = points[bool_val, :]

                    circle_bool.append(np.sum((current_points[:, 0] > corners[corner, 0]) *
                                              (current_points[:, 1] < corners[corner, 1])))

            if sum(circle_bool) >= 1:
                circle = False
                self.LogAppend.myGUI_signal.emit(
                    '[%s %s]: Arena detected as being rectangular.' % (
                        str(datetime.datetime.now().date()),
                        str(datetime.datetime.now().time())[:8]))
            else:
                circle = True
                self.LogAppend.myGUI_signal.emit(
                    '[%s %s]: Arena detected as being circular.' % (
                        str(datetime.datetime.now().date()),
                        str(datetime.datetime.now().time())[:8]))

        coverage_figure = plt.figure()
        ax_coverage = coverage_figure.add_subplot(111)
        # mng = plt.get_current_fig_manager()
        # mng.resize(*mng.window.maxsize())

        if not circle:
            # if there is data in the corners it is a square

            # find the average heigh tand average width

            bins = np.linspace(np.amin(posx), np.amax(posx), 20)
            bin_edges = np.hstack((bins[0:-1].reshape((len(bins[0:-1]), 1)),
                                   bins[1:].reshape((len(bins[1:]), 1))))

            # graphing rectangle representing the arena border

            rectangle_points = np.array([[np.amin(posx), np.amax(posy)],  # NW
                                         [np.amax(posx), np.amax(posy)],  # NE
                                         [np.amax(posx), np.amin(posy)],  # SE
                                         [np.amin(posx), np.amin(posy)],  # SW
                                         [np.amin(posx), np.amax(posy)]  # NW
                                         ])

            self.border = ax_coverage.plot(rectangle_points[:, 0], rectangle_points[:, 1], 'b', lw=plot_linewidth / 10)
            ax_coverage.plot(posx, posy, 'r-', lw=plot_linewidth)

            xlim_values = [min([dimensions[0], np.amin(rectangle_points[:, 0])]) - 0.5,
                                  max([dimensions[1], np.amax(rectangle_points[:, 0])]) + 0.5]
            ylim_values = [min([dimensions[2], np.amin(rectangle_points[:, 1])]) - 0.5,
                                  max([dimensions[3], np.amax(rectangle_points[:, 1])]) + 0.5]



            total_area_cm2 = (np.abs(np.amin(posx)) +
                              np.amax(posx)) * (np.abs(np.amin(posy) +
                                                       np.amax(posy)))  # Length * Width
        else:
            # there were no position values in the corner, must be a circle

            bins = np.linspace(np.amin(posx), np.amax(posx), 50)

            bin_edges = np.hstack((bins[0:-1].reshape((len(bins[0:-1]), 1)),
                                   bins[1:].reshape((len(bins[1:]), 1))))

            radii = np.array([np.abs(np.amin(posx)), np.amax(posx),
                              np.abs(np.amin(posy)), np.amax(posy)])

            for bin_value in range(len(bin_edges)):
                bin_bool = (posx >= bin_edges[bin_value, 0]) * (posx < bin_edges[bin_value, 1])

                if sum(bin_bool) == 0:
                    # no points in this bin
                    continue

                posx_bin = posx[bin_bool]
                posy_bin = posy[bin_bool]

                max_val = np.amax(posy_bin)
                max_i = np.where(posy_bin == max_val)[0][0]

                min_val = np.amin(posy_bin)
                min_i = np.where(posy_bin == min_val)[0][0]

                append_radii = np.array([np.sqrt(max_val ** 2 + posx_bin[max_i] ** 2),
                                         np.sqrt(min_val ** 2 + posx_bin[min_i] ** 2)])

                radii = np.concatenate((radii, append_radii))

            # equivalent to ang = 0:0.01:4*pi in matlab
            step = 0.001
            ang = np.arange(np.round((4 * np.pi + step) / step)) / (1 / step)

            xp, yp = circle_vals(0, 0, 2 * np.amax(radii), ang)

            self.border = ax_coverage.plot(xp, yp, 'b', lw=plot_linewidth / 10)
            ax_coverage.plot(posx, posy, 'r', lw=plot_linewidth)

            xlim_values = [min([dimensions[0], np.amin(xp)]) - 0.5,
                                  max([dimensions[1], np.amax(xp)]) + 0.5]
            ylim_values = [min([dimensions[2], np.amin(yp)]) - 0.5,
                                  max([dimensions[3], np.amax(yp)]) + 0.5]

            total_area_cm2 = np.pi * (np.amax(radii) ** 2)  # area = pr^2

        ax_coverage.set_xlim(xlim_values)
        ax_coverage.set_ylim(ylim_values)

        cover_png_total = os.path.join(save_figures_directory, '%s_total.png' % session)
        ax_coverage.axis('off')
        # coverage_figure.show()
        coverage_figure.savefig(cover_png_total, bbox_inches='tight', dpi=200)  # saving figure with the arena trace
        plt.close(coverage_figure)
        # now reload the image so that we can use edge detection and regionprops to calculate the area

        # finding the total area
        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: Finding the total area of the arena!' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8]))

        RGBA = misc.imread(cover_png_total)
        try:
            RGB = color.rgba2rgb(RGBA)
        except ValueError:
            RGB = RGBA

        # plt.imshow(RGB)
        # plt.axis('off')

        I = rgb2gray(RGB)

        I = np.round(I).astype('int32')

        # create a binary gradient mask of image
        BWs_x = ndimage.sobel(I, 0)  # horizontal derivative
        BWs_y = ndimage.sobel(I, 1)  # vertical derivative
        BWs = numpy.hypot(BWs_x, BWs_y)  # magnitude

        BWs *= 255.0 / numpy.amax(BWs)  # normalize (Q&D)

        # create a dilated gradient mask
        BWsdil = ndimage.morphology.binary_dilation(BWs)  # used to use this, but found it unnecessary

        BWdfill = ndimage.morphology.binary_fill_holes(BWs)

        area = []

        label_img = skimage.measure.label(BWdfill)
        regions = skimage.measure.regionprops(label_img)
        for prop in regions:
            area.append(prop.area)
            bbox = prop.bbox  # [miny, minx, maxy, maxx ]
            dy_pixels = bbox[2] - bbox[0]
            pixels_per_cm = dy_pixels / (np.amax(posy) - np.amin(posy))

        # bwsfill_fig = plt.figure(figsize=(10, 10))
        # bwsfill_fig_ax = bwsfill_fig.add_subplot(111)
        # bwsfill_fig_ax.imshow(BWdfill, cmap=plt.cm.gray)

        total_area = max(area)

        '''
        # now we will remove the border trace, and use edge detection to calculate the travelled area

        self.border[0].remove()  # remove border
        cover_png = os.path.join(save_figures_directory, '%s_coverage.png' % session)
        coverage_figure.savefig(cover_png, bbox_inches='tight', dpi=200)  # save figure without border

        # reading in the positions without the arena trace
        RGBA = misc.imread(cover_png)
        try:
            RGB = color.rgba2rgb(RGBA)
        except ValueError:
            RGB = RGBA

        I = rgb2gray(RGB)
        I = np.round(I).astype('int32')

        if np.amax(I) <= 1:
            # then the image was saved from numpy
            BWdfill = I < 1
        else:
            BWdfill = I < 255

        # finding the contours of the path so we can find the area

        if debug_figs:
            bwsfill_fig2 = plt.figure(figsize=(10, 10))
            plt.imshow(BWdfill, cmap=plt.cm.gray)
            bwsfill_fig2.show()

            # set up the 'FilledImage' bit of regionprops.
            #filledI = np.zeros(BWdfill.shape[0:2]).astype('uint8')
            # set up the 'ConvexImage' bit of regionprops.
            #a = cv2.drawContours(filledI, contours, -1, (255, 0, 0), 3)
            #contour_no_border = plt.figure()
            #plt.imshow(a, cmap=plt.cm.gray)

            coverage_fill_figure = plt.figure()
            ax_coverage_fill = coverage_fill_figure.add_subplot(111)
            #mng = plt.get_current_fig_manager()
            #mng.resize(*mng.window.maxsize())
            ax_coverage_fill.imshow(BWdfill, cmap=plt.cm.gray)

        area = []

        label_img = skimage.measure.label(BWdfill)
        regions = skimage.measure.regionprops(label_img)
        for prop in regions:
            area.append(prop.area)

        if len(regions) > 1:
            print('Too many regions!')

        total_filled = max(area)
        '''

        # now we will break the image up into quadrants.

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: breaking up arena into quadrants!' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8]))

        # finding boundary lines
        geometry = (2, 2)  # break it up into quadrants
        labels = track_to_xylabels(posx, posy, geometry)

        quadrant_figure = plt.figure()
        ax_quadrant = quadrant_figure.add_subplot(111)
        #mng = plt.get_current_fig_manager()
        #mng.resize(*mng.window.maxsize())

        plot_map_labels(posx, posy, labels, ax=ax_quadrant)

        ax_quadrant.plot(posx, posy, "k-", alpha=0.2)

        x_ticks = np.linspace(np.min(posx), np.max(posx), geometry[0] + 1)
        y_ticks = np.linspace(np.min(posy), np.max(posy), geometry[1] + 1)

        if not circle:
            # then it's a rectangular arena
            ax_quadrant.vlines(x_ticks, np.min(posy), np.max(posy), linestyles='dashed')
            ax_quadrant.hlines(y_ticks, np.min(posx), np.max(posx), linestyles='dashed')
        else:
            # plot a circular trace representing the arena, and then plot the quadrant boundaries
            # equivalent to ang = 0:0.01:4*pi in matlab
            step = 0.001
            ang = np.arange(np.round((2 * np.pi + step) / step)) / (1 / step)
            xp, yp = circle_vals(0, 0, 2 * np.amax(radii), ang)
            ax_quadrant.plot(xp, yp, 'k--')
            ax_quadrant.vlines(x_ticks[1:-1], np.min(yp), np.max(yp), linestyles='dashed')
            ax_quadrant.hlines(y_ticks[1:-1], np.min(xp), np.max(xp), linestyles='dashed')

        # save this figure
        ax_quadrant.set_xlim(xlim_values)
        ax_quadrant.set_ylim(ylim_values)

        ax_quadrant.axis('off')
        quadrant_figurename = os.path.join(save_figures_directory, '%s_quadrants.png' % session)

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: saving png of quadrants at the following location: %s!' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8], quadrant_figurename))

        quadrant_figure.savefig(quadrant_figurename, bbox_inches='tight', dpi=200)
        plt.close(quadrant_figure)
        label_data = []
        # iterate through each of the quadrants
        label_data_order = []

        max_label = np.amax(labels)

        if geometry == (2, 2):
            number_cols = np.int(np.floor(max_label))  # number of columns
            number_rows = np.int((max_label - number_cols) * 10)  # number of columns

        total_filled = 0

        quadrant_number = 1

        for label in np.unique(labels):

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: analyzing quadrant %d!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8], quadrant_number))
            quadrant_number += 1

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: analyzing amount of time spent within the quadrant!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))

            label_dictionary = {}

            label_bool = np.where(labels == label)[0]

            Fs_pos = 50  # Hz

            consecutive_points = find_consec(label_bool)

            label_dictionary['time'] = len(label_bool) / Fs_pos  # the amount of time in seconds within this sector

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: analyzing the number of times the animal entered this quadrant!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))

            label_dictionary['n_entered'] = len(
                consecutive_points)  # the amount of times that the animal entered the sector

            v_total = np.array([])

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: analyzing the average (and standard deviation) of mouse velocity within this quadrant!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))

            for run in consecutive_points:

                runx = posx[run]
                runy = posy[run]
                runt = post[run]

                if len(runx) <= 1:
                    continue

                v = speed2D(runx, runy, runt)  # finds the speed of the animal

                if len(v) != 0:
                    v_total = np.concatenate((v_total, v.flatten()))
                else:
                    v_total = v.flatten()

            v_avg = np.mean(v_total)  # finding the average velocity in this sector
            v_sd = np.std(v_total)  # finding the standard deviation of the sector

            label_dictionary['v_average'] = v_avg
            label_dictionary['v_std'] = v_sd

            # find the coverage

            quadrant_positions = plt.figure()
            quadrant_axis = quadrant_positions.add_subplot(111)

            quadrant_axis.plot(posx[label_bool], posy[label_bool], 'r-', lw=plot_linewidth)

            quadrant_axis.set_xlim(xlim_values)
            quadrant_axis.set_ylim(ylim_values)

            quadrant_position_filename = os.path.join(save_figures_directory, '%s_quadrant_%s.png' % (session, label))

            quadrant_axis.axis('off')
            quadrant_positions.savefig(quadrant_position_filename, bbox_inches='tight', dpi=200)
            plt.close(quadrant_positions)
            # finding the quadrant area

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: analyzing coverage within the quadrant!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8]))

            RGBA = misc.imread(quadrant_position_filename)
            try:
                RGB = color.rgba2rgb(RGBA)
            except ValueError:
                RGB = RGBA

            I = rgb2gray(RGB)
            I = np.round(I).astype('int32')

            if np.amax(I) <= 1:
                # then the image was saved from numpy
                BWdfill = I < 1
            else:
                BWdfill = I < 255

            if debug_figs:
                coverage_fill_figure = plt.figure()
                ax_coverage_fill = coverage_fill_figure.add_subplot(111)
                #mng = plt.get_current_fig_manager()
                #mng.resize(*mng.window.maxsize())
                ax_coverage_fill.imshow(BWdfill, cmap=plt.cm.gray)

            area = []

            label_img = skimage.measure.label(BWdfill)
            regions = skimage.measure.regionprops(label_img)
            for prop in regions:
                area.append(prop.area)

            quadrant_area = max(area)
            label_dictionary['quadrant_travelled'] = quadrant_area / (pixels_per_cm**2)  # cm**2
            label_dictionary['quadrant_perc_total'] = 100*quadrant_area/(total_area/(number_cols*number_rows))  # %
            label_dictionary['quadrant_perc_travelled'] = quadrant_area # this is just a placeholder, we will modify this value after we get the total filled value

            total_filled += quadrant_area

            label_data.append(label_dictionary)
            label_data_order.append(label)

        for i, values in enumerate(label_data):
            # now convert the quadrant_perc_travelled into a percentage
            label_data[i]['quadrant_perc_travelled'] *= (100/total_filled)  # %

        # save all values in an excel sheet

        header = ['Quadrant:', 'Time(sec):', 'N Entered:', 'Velocity Avg.(cm/s):', 'Velocity SDev.(cm/s):',
                  'Quadrant Coverage(cm*2):', 'Percent Quadrant Covered(%):', 'Percent of Animal Coverage(%):']

        # create a new worksheet

        self.LogAppend.myGUI_signal.emit(
            '[%s %s]: creating excel file!' % (
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time())[:8]))

        wb = Workbook()
        active_sheets = wb.sheetnames
        ws = wb[active_sheets[0]]
        ws.append(header)

        if geometry == (2, 2):

            # labels = col.row
            write_index = 0

            label_names = ['Upper Left', 'Upper Right', 'Lower Left', 'Lower Right']
            for row in range(number_rows, 0, -1):
                for col in range(1, number_cols + 1):

                    label = col + row/10
                    label_index = label_data_order.index(label)
                    value_dictionary = label_data[label_index]

                    values = [label_names[write_index], value_dictionary['time'], value_dictionary['n_entered'],
                              value_dictionary['v_average'], value_dictionary['v_std'],
                              value_dictionary['quadrant_travelled'], value_dictionary['quadrant_perc_total'],
                              value_dictionary['quadrant_perc_travelled']]

                    write_index += 1
                    ws.append(values)

            img = Image(quadrant_figurename)
            ws.add_image(img, 'A10')

            wb.save(excel_filename)

            self.LogAppend.myGUI_signal.emit(
                '[%s %s]: quadrant analysis excel file saved at the following location: %s!' % (
                    str(datetime.datetime.now().date()),
                    str(datetime.datetime.now().time())[:8], excel_filename))

        else:
            print('Have not coded with this geometry layout yet')

        os.remove(cover_png_total)

        for label in np.unique(labels):
            quadrant_position_filename = os.path.join(save_figures_directory, '%s_quadrant_%s.png' % (session, label))
            os.remove(quadrant_position_filename)

    self.LogAppend.myGUI_signal.emit(
        '[%s %s]: finished!' % (
            str(datetime.datetime.now().date()),
            str(datetime.datetime.now().time())[:8]))


def track_to_xylabels(x_pos, y_pos, geometry):
    """
    returns the list of labels for each datapoint

     For example: 2x2 grid (geometry = (2,2))

              y=100 -------------------------------------------
                    |                    |                     |
                    |                    |                     |
                    |   Label = 1.2      |   Label = 2.2       |    y
                    |                    |                     |    |
                    |                    |                     |    a
               y=50 -------------------------------------------     x
                    |                    |                     |    i
                    |                    |                     |    s
                    |  Label = 1.1       |   Label = 2.1       |
                    |                    |                     |
                    |                    |                     |
                y=0 -------------------------------------------|
                    x=0                 x=50                  x=100
                                x-axis

    """

    x_ticks = np.linspace(np.min(x_pos), np.max(x_pos), geometry[0] + 1)[:-1]
    y_ticks = np.linspace(np.min(y_pos), np.max(y_pos), geometry[1] + 1)[:-1]

    x_lab = np.sum([x_pos >= t for t in x_ticks], 0)
    y_lab = np.sum([y_pos >= t for t in y_ticks], 0)

    return 1. * x_lab + y_lab / 10


def xylabels_to_track(x_pos, y_pos, geometry, labels):
    x_ticks = np.linspace(np.min(x_pos), np.max(x_pos), geometry[0] + 1)[:-1]
    y_ticks = np.linspace(np.min(y_pos), np.max(y_pos), geometry[1] + 1)[:-1]

    xx = np.r_[[int(l) - 1 for l in labels]] * np.diff(x_ticks)[0] + np.diff(x_ticks)[0] / 2 + np.min(x_pos)
    yy = np.r_[[(l - int(l)) * 100 - 1 for l in labels]] * np.diff(y_ticks)[0] + np.diff(y_ticks)[0] / 2 + np.min(y_pos)

    return xx, yy


def plot_map_labels(x_pos, y_pos, labels, ax=None, **args):
    if ax is None:
        fig = plt.figure()
        ax = fig.add_subplot(111)
    colors = plt.cm.rainbow(np.linspace(0, 1, len(np.unique(labels))))
    for l, c in zip(np.unique(labels), colors):
        ax.scatter(x_pos[labels == l], y_pos[labels == l], color=c, **args)

    return ax


def find_consec(data):
    '''finds the consecutive numbers and outputs as a list'''
    consecutive_values = []  # a list for the output
    current_consecutive = [data[0]]

    if len(data) == 1:
        return [[data[0]]]

    for index in range(1, len(data)):

        if data[index] == data[index - 1] + 1:
            current_consecutive.append(data[index])

            if index == len(data) - 1:
                consecutive_values.append(current_consecutive)

        else:
            consecutive_values.append(current_consecutive)
            current_consecutive = [data[index]]

            if index == len(data) - 1:
                consecutive_values.append(current_consecutive)
    return consecutive_values


def circle_vals(x, y, d, ang):
    """x and y are the coordinates of the center o the circle with radius,r"""
    r = d / 2

    xp = np.multiply(r, np.cos(ang))
    yp = np.multiply(r, np.sin(ang))

    xp = np.add(xp, x)
    yp = np.add(yp, y)

    return xp.reshape((len(xp), 1)), yp.reshape((len(yp), 1))


def rgb2gray(rgb):
    r, g, b = rgb[:, :, 0], rgb[:, :, 1], rgb[:, :, 2]
    gray = 0.2989 * r + 0.5870 * g + 0.1140 * b

    return gray


def flood_fill(test_array, h_max=255):
    input_array = np.copy(test_array)
    el = scipy.ndimage.generate_binary_structure(2, 2).astype(np.int)
    inside_mask = scipy.ndimage.binary_erosion(~np.isnan(input_array), structure=el)
    output_array = np.copy(input_array)
    output_array[inside_mask] = h_max
    output_old_array = np.copy(input_array)
    output_old_array.fill(0)
    el = scipy.ndimage.generate_binary_structure(2, 1).astype(np.int)
    while not np.array_equal(output_old_array, output_array):
        output_old_array = np.copy(output_array)
        output_array = np.maximum(input_array, scipy.ndimage.grey_erosion(output_array, size=(3, 3), footprint=el))
    return output_array


class Window(QtGui.QWidget):  # defines the window class (main window)
    '''The function that will serve to create the main window of the GUI'''

    def __init__(self):  # initializes the main window
        super(Window, self).__init__()
        # self.setGeometry(50, 50, 500, 300)
        pass


def run():
    app = QtGui.QApplication(sys.argv)

    set_directory = 'C:\\Users\\Geoffrey Barrett\\Desktop\\B6-1M 20160430'
    window = Window()
    QuadrantAnalysis(window, set_directory, 'DarkRoom')

    sys.exit(app.exec_())  # prevents the window from immediately exiting out


if __name__ == '__main__':
    run()  # the command that calls run()